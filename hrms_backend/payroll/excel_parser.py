"""hrms_backend /payroll/excel_parser.py"""
# C:\Users\ashwi\Desktop\Replica_caoas_Backup\hrms_backend\payroll\excel_parser.py

import re
from decimal import Decimal, InvalidOperation

from django.db import IntegrityError, transaction
from django.utils import timezone
from openpyxl import load_workbook

from .models import Employee, EmployeeSalaryStructure
from .upload_validators import ExcelTooLargeError, enforce_max_rows


class ExcelParseError(Exception):
    pass


class PayrollRowValidationError(Exception):
    def __init__(self, errors: list[dict]):
        self.errors = errors
        super().__init__(f"{len(errors)} rows have errors")


def _normalize_header(value) -> str:
    """
    Collapses internal whitespace/newlines/tabs to single spaces and
    strips — real-world excel headers are frequently exported with
    inconsistent line breaks or double spaces (seen firsthand in the
    actual upload sheet), so exact-string matching against a fixed
    "\\n"-embedded literal is too fragile. Both COLUMN_MAPPING keys and
    the sheet's own header row are normalized the same way before
    comparison.
    """
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


# This is the REAL monthly upload sheet's column set — 16 columns of
# attendance + uploaded allowances/deductions. Everything else (basic_da,
# hra, basic_for_pf, gross_salary, earned_salary, epf, vpf, professional_tax,
# net_salary, total_deductions) used to be required here but is now
# computed server-side in calculations.py from EmployeeSalaryStructure,
# since the real sheet never carried those columns and every upload
# against it was failing at the "template is missing columns" check.
# Night Shift removed (soft-removed, DB column kept for history) — its pay
# is now expected to be folded into another existing field before upload.
VARIABLE_COLUMN_MAPPING = {
    "Actual Working Days":  "actual_working_days",
    "Extra Working Days":   "extra_working_days",
    "Paid Leave Days":      "paid_leave_days",
    "LOP":                  "lop_days",
    "Leave Travel Allowance": "lta",
    "Special Allowance":     "special_allowance",
    "NPS Allowance":         "nps_allowance_earned",
    "commission/other allowance/Retention Bonus": "commission_other",
    "Arrears":               "arrears",
    "Reimbursements":        "reimbursements",
    "TDS":                   "tds",
    "VPF Arrears":            "vpf_arrears",
    "NPS Deduction - Arrears": "nps_deduction_arrears",
    "Loan Deduction":         "loan_deduction",
    "LWF":                    "lwf",
    "Other deduction":        "other_deduction",
}
# Normalize the mapping's own keys once, at import time.
VARIABLE_COLUMN_MAPPING = {_normalize_header(k): v for k, v in VARIABLE_COLUMN_MAPPING.items()}

EMPLOYEE_CODE_HEADER = "Employee Code"

DECIMAL_FIELDS = {
    "paid_leave_days", "lop_days", "extra_working_days",
    "lta", "special_allowance", "nps_allowance_earned", "commission_other",
    "arrears", "reimbursements", "tds", "vpf_arrears", "nps_deduction_arrears",
    "loan_deduction", "lwf", "other_deduction",
}

INTEGER_FIELDS = {"actual_working_days"}


def _is_empty_row(values) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)


def _json_safe(value):
    if isinstance(value, Decimal):
        return str(value)
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _to_decimal(value, field_name: str, errors: list[str]) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    try:
        amount = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError):
        errors.append(f"{field_name} must be numeric")
        return Decimal("0")
    if amount < 0:
        errors.append(f"{field_name} must be greater than or equal to 0")
    return amount.quantize(Decimal("0.01"))


def _to_int(value, field_name: str, errors: list[str]) -> int:
    amount = _to_decimal(value, field_name, errors)
    if amount != amount.to_integral_value():
        errors.append(f"{field_name} must be a whole number")
    return int(amount)


def _header_indexes(headers: list[str]) -> dict[str, int]:
    indexes = {}
    for idx, raw_header in enumerate(headers, start=1):
        header = _normalize_header(raw_header)
        if header == EMPLOYEE_CODE_HEADER:
            indexes["employee_code"] = idx
        elif header in VARIABLE_COLUMN_MAPPING:
            indexes[VARIABLE_COLUMN_MAPPING[header]] = idx
    return indexes


def parse_payroll_excel(file, client) -> tuple[list[dict], list[str]]:
    try:
        workbook = load_workbook(file, data_only=True)
    except Exception as exc:
        raise ExcelParseError(f"Unable to read Excel file: {exc}") from exc

    worksheet = workbook.active
    try:
        enforce_max_rows(worksheet)
    except ExcelTooLargeError as exc:
        raise ExcelParseError(str(exc)) from exc

    header_row_number = 5
    headers = [cell.value for cell in worksheet[header_row_number]]
    indexes = _header_indexes(headers)

    if "employee_code" not in indexes:
        raise ExcelParseError("Header row not found or Employee Code column is missing.")

    missing = [field for field in VARIABLE_COLUMN_MAPPING.values() if field not in indexes]
    if missing:
        raise ExcelParseError(f"Payroll template is missing columns: {', '.join(missing)}")

    employee_codes = set()
    row_payloads = []
    row_errors = []
    warnings = []

    for row_number in range(6, worksheet.max_row + 1):
        cells = list(worksheet[row_number])
        values = [cell.value for cell in cells]
        if _is_empty_row(values):
            continue

        raw_row = {
            _normalize_header(headers[idx - 1]) or f"Column {idx}": _json_safe(values[idx - 1])
            for idx in range(1, len(values) + 1)
        }
        errors = []
        employee_code_raw = worksheet.cell(row=row_number, column=indexes["employee_code"]).value
        employee_code = str(employee_code_raw).strip() if employee_code_raw is not None else ""

        if not employee_code:
            errors.append("Employee Code is required")
        else:
            employee_codes.add(employee_code)

        parsed = {
            "row_number": row_number,
            "employee_code": employee_code,
            "raw_row_data": raw_row,
        }

        for field_name, col_index in indexes.items():
            if field_name == "employee_code":
                continue
            value = worksheet.cell(row=row_number, column=col_index).value
            if field_name in INTEGER_FIELDS:
                parsed[field_name] = _to_int(value, field_name, errors)
            elif field_name in DECIMAL_FIELDS:
                parsed[field_name] = _to_decimal(value, field_name, errors)

        if errors:
            row_errors.append({"row": row_number, "employee_code": employee_code, "errors": errors})
        row_payloads.append(parsed)

    # Employee is scoped to the selected client now (payroll.Employee.client)
    # — a row whose employee_code doesn't belong to THIS client is rejected,
    # even if that code exists for a different client, so batches never
    # cross-contaminate between clients. `status="active"` is Employee's
    # equivalent of an is_active flag (see models.py STATUS_CHOICES).
    employees_by_code = {
        employee.employee_code: employee
        for employee in Employee.objects.filter(
            client=client, employee_code__in=employee_codes, status=Employee.STATUS_ACTIVE,
        )
    }

    parsed_rows = []
    for parsed in row_payloads:
        employee_code = parsed["employee_code"]
        if employee_code and employee_code not in employees_by_code:
            row_errors.append(
                {
                    "row": parsed["row_number"],
                    "employee_code": employee_code,
                    "errors": [
                        f"Employee Code {employee_code} not found for this client. "
                        "Import this employee via the employee-master upload before "
                        "uploading payroll."
                    ],
                }
            )
            continue
        parsed["employee"] = employees_by_code.get(employee_code)
        parsed_rows.append(parsed)

    if row_errors:
        raise PayrollRowValidationError(row_errors)

    return parsed_rows, warnings


# ── Employee master (bulk import) ──────────────────────────────────────────
# Separate upload flow from the monthly payroll batch above: this
# creates/updates payroll.Employee rows themselves for one client, so the
# monthly upload above has employees to match against. Same "reject bad
# rows with a clear error, never silently skip" policy as the batch parser.

EMPLOYEE_MASTER_COLUMN_MAPPING = {
    "Employee Code": "employee_code",
    "First Name": "first_name",
    "Last Name": "last_name",
    "Email": "email",
    "PAN Number": "pan_number",
    "Department": "department",
    "Position": "position",
    "Hire Date": "hire_date",
    "CTC": "ctc",
    "Status": "status",
    # Not an Employee field — used only to feed pf_opted into
    # EmployeeSalaryStructure.build_from_ctc() below for rows that get
    # an auto-created initial structure. Doesn't touch Employee itself.
    "PF Applicable": "pf_applicable",
}
EMPLOYEE_MASTER_COLUMN_MAPPING = {
    _normalize_header(k): v for k, v in EMPLOYEE_MASTER_COLUMN_MAPPING.items()
}


class EmployeeImportError(Exception):
    def __init__(self, errors: list[dict]):
        self.errors = errors
        super().__init__(f"{len(errors)} rows have errors")


def _to_date(value, field_name: str, errors: list[str]):
    if value is None or value == "":
        return None
    if hasattr(value, "date"):
        return value.date() if hasattr(value, "hour") else value
    errors.append(f"{field_name} must be a valid date")
    return None


def parse_employee_master_excel(file, client, created_by=None) -> tuple[int, int, list[str]]:
    """
    Creates/updates payroll.Employee rows for `client` from an uploaded
    Excel file. Matches existing rows by employee_code WITHIN this
    client (create if new, update if the code already exists for this
    client — a code can be reused across different clients). Returns
    (created_count, updated_count, warnings). Raises EmployeeImportError
    with row-level detail on any bad row — nothing is partially applied.

    This import is for onboarding: for any row with a CTC value where the
    employee does NOT already have a salary structure on file, an initial
    EmployeeSalaryStructure is auto-derived from that CTC via
    EmployeeSalaryStructure.build_from_ctc() (same math as
    SalaryStructureModal.jsx). The optional "PF Applicable" column
    (yes/no, defaults to yes if blank) controls pf_opted for that
    auto-created structure. Employees who already have a structure are
    left untouched here — re-importing only updates their plain Employee
    fields; structure changes after onboarding go through the normal
    Edit Salary Structure flow, not this bulk import.
    """
    try:
        workbook = load_workbook(file, data_only=True)
    except Exception as exc:
        raise ExcelParseError(f"Unable to read Excel file: {exc}") from exc

    worksheet = workbook.active
    try:
        enforce_max_rows(worksheet)
    except ExcelTooLargeError as exc:
        raise ExcelParseError(str(exc)) from exc

    headers = [cell.value for cell in worksheet[1]]
    indexes = {}
    for idx, raw_header in enumerate(headers, start=1):
        header = _normalize_header(raw_header)
        if header in EMPLOYEE_MASTER_COLUMN_MAPPING:
            indexes[EMPLOYEE_MASTER_COLUMN_MAPPING[header]] = idx

    if "employee_code" not in indexes:
        raise ExcelParseError("Header row not found or Employee Code column is missing.")

    row_errors = []
    parsed_rows = []
    warnings = []

    for row_number in range(2, worksheet.max_row + 1):
        values = [cell.value for cell in worksheet[row_number]]
        if _is_empty_row(values):
            continue

        errors = []
        code_raw = worksheet.cell(row=row_number, column=indexes["employee_code"]).value
        employee_code = str(code_raw).strip() if code_raw is not None else ""
        if not employee_code:
            errors.append("Employee Code is required")

        def cell(field):
            col = indexes.get(field)
            return worksheet.cell(row=row_number, column=col).value if col else None

        first_name = str(cell("first_name") or "").strip()
        if not first_name:
            errors.append("First Name is required")

        ctc_raw = cell("ctc")
        ctc = None
        if ctc_raw not in (None, ""):
            ctc = _to_decimal(ctc_raw, "CTC", errors)

        hire_date = _to_date(cell("hire_date"), "Hire Date", errors)

        status_raw = str(cell("status") or Employee.STATUS_ACTIVE).strip().lower()
        if status_raw not in (Employee.STATUS_ACTIVE, Employee.STATUS_INACTIVE):
            errors.append(f"Status must be one of: {Employee.STATUS_ACTIVE}, {Employee.STATUS_INACTIVE}")

        pf_raw = cell("pf_applicable")
        pf_text = str(pf_raw).strip().lower() if pf_raw not in (None, "") else "yes"
        if pf_text not in ("yes", "no"):
            errors.append("PF Applicable must be 'yes' or 'no'")
        pf_opted = pf_text == "yes"

        if errors:
            row_errors.append({"row": row_number, "employee_code": employee_code, "errors": errors})
            continue

        parsed_rows.append({
            "row_number": row_number,
            "employee_code": employee_code,
            "first_name": first_name,
            "last_name": str(cell("last_name") or "").strip(),
            "email": str(cell("email") or "").strip(),
            "pan_number": str(cell("pan_number") or "").strip(),
            "department": str(cell("department") or "").strip(),
            "position": str(cell("position") or "").strip(),
            "pf_opted": pf_opted,
            "hire_date": hire_date,
            "ctc": ctc,
            "status": status_raw,
        })

    if row_errors:
        raise EmployeeImportError(row_errors)

    existing_by_code = {
        e.employee_code: e for e in Employee.objects.filter(client=client)
    }

    created_count = 0
    updated_count = 0
    duplicate_errors = []
    for row in parsed_rows:
        row_number = row.pop("row_number")
        code = row.pop("employee_code")
        pf_opted = row.pop("pf_opted")
        ctc = row.get("ctc")
        employee = existing_by_code.get(code)
        is_new = employee is None
        try:
            with transaction.atomic():
                if employee:
                    for field, value in row.items():
                        setattr(employee, field, value)
                    employee.save()
                else:
                    employee = Employee.objects.create(employee_code=code, client=client, **row)

                if ctc not in (None, "") and not employee.salary_structures.exists():
                    structure_fields = EmployeeSalaryStructure.build_from_ctc(ctc, pf_opted=pf_opted)
                    EmployeeSalaryStructure.objects.create(
                        employee=employee,
                        effective_from=employee.hire_date or timezone.now().date(),
                        created_by=created_by,
                        change_reason="Initial structure — Excel import",
                        **structure_fields,
                    )
        except IntegrityError:
            duplicate_errors.append({
                "row": row_number,
                "employee_code": code,
                "errors": ["Employee Code already exists"],
            })
            continue

        existing_by_code[code] = employee
        if is_new:
            created_count += 1
        else:
            updated_count += 1

    if duplicate_errors:
        raise EmployeeImportError(duplicate_errors)

    return created_count, updated_count, warnings
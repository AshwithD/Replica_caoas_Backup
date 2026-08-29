"""hrms_backend/payroll/views.py"""

import csv
from calendar import month_name
from io import BytesIO, StringIO
from pathlib import Path

from decimal import Decimal
from datetime import date
from unittest.mock import patch

from django.db import models, transaction
from django.http import FileResponse, Http404, HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

from django.contrib.auth import get_user_model
User = get_user_model()
from .mixins import AuditViewMixin
from .permissions import IsAdminOrManagerRole
from .email_service import send_payslip_email
from .tasks import send_bulk_payslip_emails
from celery.result import AsyncResult
from django.core.mail import get_connection
import calendar

from .calculations import calculate_payslip_fields, is_in_probation
from .excel_parser import ExcelParseError, PayrollRowValidationError, parse_payroll_excel
from .models import (
    Client, CompOffAdjustment, EmailLog, Employee, EmployeeSalaryStructure, LeaveAdjustment,
    OnHoldAdjustment, PayrollBatch, PayslipRecord, PayslipRecordEdit, SalaryAdvance,
    SalaryAdvanceAdjustment, get_latest_salary_structure,
)
from .pdf_generator import DESIGN_MODULES, generate_payslip_pdf
from .serializers import (
    EmailLogSerializer,
    PAYSLIP_EDITABLE_FIELDS,
    PayrollBatchListSerializer,
    PayrollBatchSerializer,
    PayslipRecordEditSerializer,
    PayslipRecordSerializer,
)


def _auto_generate_pending_emi_recoveries(employee, generated_by):
    """
    For every active SalaryAdvance plan belonging to this employee
    (months_recovered < tenure_months), auto-create this month's
    -emi_amount recovery SalaryAdvanceAdjustment if one isn't already
    waiting to be applied — so recovery happens automatically each
    payroll cycle with no manual monthly adjustment required.

    Two guards keep this safe to call every time a batch is generated:
    - Skipped entirely if the plan's own disbursement adjustment hasn't
      been applied to a batch yet (i.e. the advance was just created this
      same cycle) — recovery starts the cycle AFTER disbursement, not the
      same one, so an employee isn't given and docked an advance in the
      same month.
    - Skipped if this plan already has an unapplied (pending) recovery
      adjustment sitting there — prevents double-deducting if batch
      generation is ever re-run for the same month.

    The final installment (months_recovered == tenure_months - 1 going
    in) recovers whatever amount remains rather than another flat
    emi_amount, so the total recovered exactly equals total_amount
    regardless of any rounding in emi_amount.

    Newly created recovery adjustments are left pending (applied_in_record
    left null) here — they get folded into salary_advance_recovered and
    marked applied by the existing pending_advance_adjustments handling
    right below each call site, same as any other pending adjustment.
    months_recovered is only incremented once a recovery adjustment is
    actually applied — see that call site.
    """
    active_advances = (
        SalaryAdvance.objects
        .filter(employee=employee, months_recovered__lt=models.F("tenure_months"))
        .select_related("disbursement_adjustment")
    )
    for advance in active_advances:
        if advance.disbursement_adjustment.applied_in_record_id is None:
            continue
        has_pending_recovery = advance.adjustments.filter(
            applied_in_record__isnull=True, amount__lt=0,
        ).exists()
        if has_pending_recovery:
            continue

        installments_remaining = advance.tenure_months - advance.months_recovered
        if installments_remaining <= 1:
            already_recovered = advance.emi_amount * advance.months_recovered
            recovery_amount = advance.total_amount - already_recovered
        else:
            recovery_amount = advance.emi_amount

        SalaryAdvanceAdjustment.objects.create(
            employee=employee,
            amount=-recovery_amount,
            reason=(
                f"Auto-recovered EMI installment {advance.months_recovered + 1}/"
                f"{advance.tenure_months} for salary advance #{advance.id}"
            ),
            advance=advance,
            created_by=generated_by,
        )


def _current_open_record(employee):
    """
    The one PayslipRecord (if any) a freshly-created ledger adjustment
    should apply to immediately instead of sitting pending until the next
    batch upload — the employee's most recent record whose batch is still
    STATUS_UPLOADED (i.e. not yet reviewed/sent/locked; matches
    BatchReview.jsx's `ledgerLocked = batch.status !== "UPLOADED"`).
    None if there's no such open batch, in which case the adjustment is
    left pending and folds into salary_advance_opening_balance/etc. the
    normal way the next time a batch is uploaded for this employee.
    """
    return (
        PayslipRecord.objects
        .filter(employee=employee, batch__status=PayrollBatch.STATUS_UPLOADED)
        .select_related("batch", "salary_structure_snapshot", "employee")
        .order_by("-batch__year", "-batch__month")
        .first()
    )


def _apply_comp_off_or_leave_adjustment(record, field: str, amount: Decimal):
    """
    Immediately folds a new comp_off/leave adjustment into the currently
    open record: bumps the relevant opening balance, then re-derives every
    field calculate_payslip_fields would produce from it (closing balance,
    days_used, and the attendance-driven pay fields that depend on the
    settlement between the two ledgers) — same recompute this record
    already goes through when an attendance field is edited post-generation
    (see PayslipRecordViewSet.partial_update).
    """
    if field == "comp_off":
        record.comp_off_opening_balance += amount
    else:
        record.leave_opening_balance += amount

    is_probation = is_in_probation(record.employee.hire_date, record.batch.year, record.batch.month)
    row = {
        "actual_working_days": record.actual_working_days,
        "lop_days": record.lop_days,
        "extra_working_days": record.extra_working_days,
        "paid_leave_days": record.paid_leave_days,
        "lta": record.lta_upload,
        "special_allowance": record.special_allowance_upload,
        "nps_allowance_earned": record.nps_allowance_upload,
        "commission_other": record.commission_other,
        "arrears": record.arrears,
    }
    computed = calculate_payslip_fields(
        record.salary_structure_snapshot, row, record.batch.month, record.days_in_month,
        comp_off_opening_balance=record.comp_off_opening_balance,
        leave_opening_balance=record.leave_opening_balance,
        is_probation=is_probation,
    )
    for field_name, value in computed.items():
        if field_name in ("gross_salary", "comp_off_opening_balance", "leave_opening_balance"):
            continue
        setattr(record, field_name, value)
    record.save(update_fields=[
        "comp_off_opening_balance", "leave_opening_balance",
        "basic_da", "hra", "lta", "special_allowance", "nps_allowance_earned",
        "basic_for_pf", "epf", "vpf", "professional_tax",
        "comp_off_days_used", "comp_off_closing_balance",
        "leave_accrued", "leave_used", "leave_closing_balance",
        "earned_salary", "total_deductions", "net_salary", "updated_at",
    ])


def _apply_money_adjustment(record, field: str, amount: Decimal):
    """
    Immediately folds a new salary_advance/on_hold adjustment into the
    currently open record — these two ledgers don't feed back into
    proration/pay fields the way comp-off/leave do (see PayslipRecord.save,
    which is what actually derives *_closing_balance and earned/net salary
    from these fields), so a straight increment + save is enough.
    """
    if field == "salary_advance":
        if amount >= 0:
            record.salary_advance_disbursed += amount
        else:
            record.salary_advance_recovered += -amount
    else:
        if amount >= 0:
            record.on_hold_deducted += amount
        else:
            record.on_hold_released += -amount
    record.save()


class PayrollBatchViewSet(AuditViewMixin, viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsAdminOrManagerRole]
    http_method_names = ["get", "post", "patch", "delete", "head", "options"]

    def get_queryset(self):
        queryset = (
            PayrollBatch.objects.select_related("client", "uploaded_by", "reviewed_by", "sent_by")
            .prefetch_related("records")
            .order_by("-year", "-month", "-created_at")
        )
        if status_param := self.request.query_params.get("status"):
            queryset = queryset.filter(status=status_param)
        if year_param := self.request.query_params.get("year"):
            # Guard against partial/non-numeric input while the user is
            # still typing (e.g. "2" or "20") — .filter(year="20") would
            # otherwise raise a ValueError from the PositiveSmallIntegerField
            # lookup and 500 the request instead of just narrowing results.
            if year_param.isdigit():
                queryset = queryset.filter(year=int(year_param))
        return queryset

    def get_serializer_class(self):
        if self.action == "list":
            return PayrollBatchListSerializer
        return PayrollBatchSerializer

    def perform_destroy(self, instance):
        # An UPLOADED batch was never reviewed or sent — nothing about it
        # has audit value, so cancelling it should actually remove it
        # instead of leaving a "FAILED" husk behind. Anything past that
        # point (REVIEWED / SENDING / COMPLETED / already FAILED) keeps
        # the soft-delete so the audit trail survives.
        if instance.status == PayrollBatch.STATUS_UPLOADED:
            self._hard_delete(instance)
        else:
            instance.status = PayrollBatch.STATUS_FAILED
            instance.error_log = "Batch cancelled by user."
            instance.save(update_fields=["status", "error_log", "updated_at"])

    @staticmethod
    def _hard_delete(instance):
        # PayslipRecord/PayslipRecordEdit/EmailLog all use on_delete=PROTECT,
        # not CASCADE, so instance.delete() would raise ProtectedError.
        # Walk the dependency chain manually, leaf-first.
        record_ids = list(instance.records.values_list("id", flat=True))
        EmailLog.objects.filter(payslip_record_id__in=record_ids).delete()
        PayslipRecordEdit.objects.filter(payslip_record_id__in=record_ids).delete()
        instance.records.all().delete()
        instance.delete()

    @action(detail=True, methods=["post"], url_path="discard")
    def discard(self, request, pk=None):
        """
        Hard-delete cleanup for batches already stuck in FAILED from
        before this fix (or from a genuine row-processing error where
        soft-delete is still correct behaviour going forward, but the
        user wants this specific orphan gone). Only allowed on FAILED
        batches — anything REVIEWED/SENDING/COMPLETED keeps PROTECT.
        """
        batch = self.get_object()
        if batch.status != PayrollBatch.STATUS_FAILED:
            return Response(
                {"detail": "Only FAILED batches can be discarded."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        self._hard_delete(batch)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=["post"], url_path="upload")
    def upload(self, request):
        client_id = request.data.get("client_id") or request.data.get("client")
        month = request.data.get("month")
        year = request.data.get("year")
        file = request.FILES.get("file")

        if not all([client_id, month, year, file]):
            return Response(
                {"detail": "client_id, month, year, and file are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            month = int(month)
            year = int(year)
        except (TypeError, ValueError):
            return Response({"detail": "month and year must be integers."}, status=status.HTTP_400_BAD_REQUEST)
        if month < 1 or month > 12:
            return Response({"detail": "month must be between 1 and 12."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            client = Client.objects.get(pk=client_id)
        except (Client.DoesNotExist, ValueError, TypeError):
            return Response({"detail": "client_id does not match a known client."}, status=status.HTTP_400_BAD_REQUEST)

        if PayrollBatch.objects.filter(client=client, month=month, year=year).exists():
            return Response(
                {"detail": "Payroll batch already exists for this client, month, and year."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # parse_payroll_excel cross-checks each row's employee_code
            # against Employee.objects.filter(client=client, ...) — rows
            # for employees not on file for this client are rejected, not
            # silently skipped (see excel_parser.py).
            parsed_rows, warnings = parse_payroll_excel(file, client)
        except ExcelParseError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        except PayrollRowValidationError as exc:
            return Response(
                {
                    "detail": f"{len(exc.errors)} rows have errors",
                    "errors": exc.errors,
                    "total_errors": len(exc.errors),
                },
                status=status.HTTP_422_UNPROCESSABLE_ENTITY,
            )

        file.seek(0)
        days_in_month = calendar.monthrange(year, month)[1]  # NOT uploaded — the real
        # sheet has no "No of Days" column; calendar days is the correct
        # proration denominator, confirmed with the user.

        with transaction.atomic():
            batch = PayrollBatch.objects.create(
                client=client,
                month=month,
                year=year,
                source_file=file,
                total_records=0,
                error_log="\n".join(warnings),
                uploaded_by=request.user,
            )
            missing_structure_employees = []
            for row in parsed_rows:
                employee = row.pop("employee")
                row.pop("employee_code", None)
                row.pop("row_number", None)
                latest_structure = get_latest_salary_structure(employee)
                if latest_structure is None:
                    missing_structure_employees.append(employee)

                # Comp-off ledger: this month's opening balance is the
                # closing balance of this employee's most recent PRIOR
                # payslip record (by batch year/month), or 0 if this is
                # their first-ever record, PLUS any manual CompOffAdjustment
                # corrections made since that prior record (e.g. HR grants/
                # deducts days via the Employee Detail page) — see
                # apps.employees.views.EmployeeViewSet.comp_off_adjustments.
                # Read-only from the reviewer's perspective otherwise —
                # never uploaded, never manually set on the payslip itself.
                prior_record = (
                    PayslipRecord.objects
                    .filter(employee=employee)
                    .filter(models.Q(batch__year__lt=year) | models.Q(batch__year=year, batch__month__lt=month))
                    .order_by("-batch__year", "-batch__month")
                    .first()
                )
                pending_adjustments = list(
                    CompOffAdjustment.objects.filter(employee=employee, applied_in_record__isnull=True)
                )
                pending_adjustment_total = sum((a.amount for a in pending_adjustments), Decimal("0"))
                comp_off_opening_balance = (
                    (prior_record.comp_off_closing_balance if prior_record else Decimal("0"))
                    + pending_adjustment_total
                )

                # General leave ledger: same "prior closing balance + any
                # pending manual adjustment" pattern as comp-off above —
                # a pending adjustment here is typically a cashout (see
                # apps.employees.views.EmployeeViewSet.leave_adjustments).
                pending_leave_adjustments = list(
                    LeaveAdjustment.objects.filter(employee=employee, applied_in_record__isnull=True)
                )
                pending_leave_adjustment_total = sum((a.amount for a in pending_leave_adjustments), Decimal("0"))
                leave_opening_balance = (
                    (prior_record.leave_closing_balance if prior_record else Decimal("0"))
                    + pending_leave_adjustment_total
                )
                is_probation = is_in_probation(employee.hire_date, year, month)

                computed = calculate_payslip_fields(
                    latest_structure, row, month, days_in_month,
                    comp_off_opening_balance=comp_off_opening_balance,
                    leave_opening_balance=leave_opening_balance,
                    is_probation=is_probation,
                )

                # Salary Advance / On Hold ledgers: opening balance is a
                # plain carry-forward of the prior record's closing balance.
                # Pending adjustments are NOT folded into opening directly —
                # they become this month's real disbursed/recovered (or
                # deducted/released) movement, which both (a) shows up as
                # an actual earning/deduction line on this month's payslip
                # and (b) is what PayslipRecord.save() uses to (re)compute
                # the closing balance. Split by sign so a mix of both types
                # pending at once (e.g. one grant + one recovery) shows
                # correctly as both lines rather than a single net figure.
                # Auto-generate this month's EMI recovery adjustment(s), if
                # any of this employee's SalaryAdvance plans are due one —
                # see _auto_generate_pending_emi_recoveries. Must run before
                # the pending_advance_adjustments query right below so any
                # newly created recovery rows are picked up by it this same
                # cycle, same as a manual adjustment would be.
                _auto_generate_pending_emi_recoveries(employee, generated_by=request.user)

                pending_advance_adjustments = list(
                    SalaryAdvanceAdjustment.objects.filter(employee=employee, applied_in_record__isnull=True)
                )
                salary_advance_opening_balance = (
                    prior_record.salary_advance_closing_balance if prior_record else Decimal("0")
                )
                salary_advance_disbursed = sum(
                    (a.amount for a in pending_advance_adjustments if a.amount >= 0), Decimal("0")
                )
                salary_advance_recovered = sum(
                    (-a.amount for a in pending_advance_adjustments if a.amount < 0), Decimal("0")
                )

                pending_on_hold_adjustments = list(
                    OnHoldAdjustment.objects.filter(employee=employee, applied_in_record__isnull=True)
                )
                on_hold_opening_balance = (
                    prior_record.on_hold_closing_balance if prior_record else Decimal("0")
                )
                on_hold_deducted = sum(
                    (a.amount for a in pending_on_hold_adjustments if a.amount >= 0), Decimal("0")
                )
                on_hold_released = sum(
                    (-a.amount for a in pending_on_hold_adjustments if a.amount < 0), Decimal("0")
                )

                # lta_upload/special_allowance_upload/nps_allowance_upload
                # store the RAW excel figure alone (used later to
                # reconstruct `row` for the attendance-edit recompute path
                # — see the PayslipRecordEdit handling below). Captured
                # here, before the pops, since row.get(...) below would
                # otherwise see them already removed.
                lta_upload = row.get("lta", Decimal("0"))
                special_allowance_upload = row.get("special_allowance", Decimal("0"))
                nps_allowance_upload = row.get("nps_allowance_earned", Decimal("0"))

                # `computed` already carries the combined (prorated
                # structure baseline + this raw upload figure) value for
                # each of these three fields — see calculate_payslip_fields.
                # They must be popped out of `row` before the **row/**computed
                # spread below, or Python raises "got multiple values for
                # keyword argument" since both dicts would supply the same key.
                row.pop("lta", None)
                row.pop("special_allowance", None)
                row.pop("nps_allowance_earned", None)

                new_record = PayslipRecord.objects.create(
                    batch=batch,
                    employee=employee,
                    salary_structure_snapshot=latest_structure,
                    days_in_month=days_in_month,
                    salary_advance_opening_balance=salary_advance_opening_balance,
                    salary_advance_disbursed=salary_advance_disbursed,
                    salary_advance_recovered=salary_advance_recovered,
                    on_hold_opening_balance=on_hold_opening_balance,
                    on_hold_deducted=on_hold_deducted,
                    on_hold_released=on_hold_released,
                    lta_upload=lta_upload,
                    special_allowance_upload=special_allowance_upload,
                    nps_allowance_upload=nps_allowance_upload,
                    **row,
                    **computed,
                )
                if pending_adjustments:
                    CompOffAdjustment.objects.filter(
                        id__in=[a.id for a in pending_adjustments]
                    ).update(applied_in_record=new_record)
                if pending_leave_adjustments:
                    LeaveAdjustment.objects.filter(
                        id__in=[a.id for a in pending_leave_adjustments]
                    ).update(applied_in_record=new_record)
                if pending_advance_adjustments:
                    SalaryAdvanceAdjustment.objects.filter(
                        id__in=[a.id for a in pending_advance_adjustments]
                    ).update(applied_in_record=new_record)
                    # Recovery installments (amount < 0) tied to a
                    # SalaryAdvance plan have now been applied — advance
                    # months_recovered so the next batch knows this
                    # installment is done. The disbursement adjustment
                    # (amount >= 0) is excluded — it doesn't count as a
                    # recovered installment.
                    recovered_advance_ids = [
                        a.advance_id for a in pending_advance_adjustments
                        if a.advance_id is not None and a.amount < 0
                    ]
                    if recovered_advance_ids:
                        SalaryAdvance.objects.filter(id__in=recovered_advance_ids).update(
                            months_recovered=models.F("months_recovered") + 1
                        )
                if pending_on_hold_adjustments:
                    OnHoldAdjustment.objects.filter(
                        id__in=[a.id for a in pending_on_hold_adjustments]
                    ).update(applied_in_record=new_record)
            batch.total_records = len(parsed_rows)
            batch.save(update_fields=["total_records", "updated_at"])

            if missing_structure_employees:
                codes = ", ".join(e.employee_code for e in missing_structure_employees)
                users = User.objects.filter(role__in=[User.ROLE_ADMIN, User.ROLE_MANAGER])
                Notification.objects.bulk_create([
                    Notification(
                        user=user,
                        title="Salary structure missing",
                        message=(
                            f"{client.name} payroll for {month:02d}/{year}: "
                            f"{len(missing_structure_employees)} employee(s) have no salary "
                            f"structure set up ({codes}) — their pay will calculate as 0 until "
                            f"this is fixed."
                        ),
                    )
                    for user in users
                ])

        serializer = PayrollBatchSerializer(batch, context={"request": request})
        return Response({"batch": serializer.data, "warnings": warnings}, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["post"], url_path="mark-reviewed")
    def mark_reviewed(self, request, pk=None):
        batch = self.get_object()
        results = []
        with transaction.atomic():
            for record in batch.records.select_related("employee", "batch__client"):
                pdf_path = generate_payslip_pdf(record)
                results.append({"record": record.id, "pdf_path": pdf_path})
            batch.status = PayrollBatch.STATUS_REVIEWED
            batch.reviewed_by = request.user
            batch.reviewed_at = timezone.now()
            batch.save(update_fields=["status", "reviewed_by", "reviewed_at", "updated_at"])

        return Response({"batch": PayrollBatchSerializer(batch, context={"request": request}).data, "results": results})

    @action(detail=True, methods=["post"], url_path="send-emails")
    def send_emails(self, request, pk=None):
        batch = self.get_object()
        if batch.status not in [PayrollBatch.STATUS_REVIEWED, PayrollBatch.STATUS_SENDING, PayrollBatch.STATUS_FAILED]:
            return Response(
                {"detail": "Batch must be REVIEWED, SENDING, or FAILED before emails can be sent."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # The actual sending (and the status/notification bookkeeping that
        # goes with it) now runs in the background via Celery — see
        # apps/payroll/tasks.py. A 50-employee batch used to block this
        # request for several minutes, which risked proxy/browser timeouts;
        # this returns immediately instead. Poll the batch's own status
        # field (STATUS_SENDING -> STATUS_COMPLETED/STATUS_FAILED) or the
        # Celery task id below to track progress.
        task = send_bulk_payslip_emails.delay(batch.id, request.user.id)

        return Response(
            {"detail": "Payslip emails are being sent in the background.", "task_id": task.id},
            status=status.HTTP_202_ACCEPTED,
        )

    @action(detail=True, methods=["get"], url_path="send-emails-status/(?P<task_id>[^/.]+)")
    def send_emails_status(self, request, pk=None, task_id=None):
        """
        Frontend polls this after send-emails dispatches a task, until
        state is SUCCESS or FAILURE — result (once SUCCESS) is the exact
        same {batch_status, total_sent, total_failed, results} shape the
        send-emails endpoint used to return synchronously, so
        SendResultsModal doesn't need to change how it reads the data.
        """
        result = AsyncResult(task_id)
        payload = {"state": result.state}
        if result.state == "SUCCESS":
            payload["result"] = result.result
        elif result.state == "FAILURE":
            payload["error"] = str(result.result)
        return Response(payload)

    @action(detail=True, methods=["get"], url_path="export-csv")
    def export_csv(self, request, pk=None):
        batch = self.get_object()
        buffer = StringIO()
        writer = csv.writer(buffer)
        writer.writerow(["Employee Code", "Name", "Email", "Gross Salary", "Net Salary", "Status", "Sent At"])
        for record in batch.records.select_related("employee").prefetch_related("email_logs"):
            latest_log = record.email_logs.order_by("-sent_at").first()
            writer.writerow(
                [
                    record.employee.employee_code,
                    record.employee.full_name,
                    record.employee.email,
                    record.gross_salary,
                    record.net_salary,
                    record.status,
                    latest_log.sent_at.isoformat() if latest_log else "",
                ]
            )
        response = HttpResponse(buffer.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="payroll_{batch.client_id}_{batch.year}_{batch.month:02d}.csv"'
        return response

    @action(detail=True, methods=["get"], url_path="export-detailed-xlsx")
    def export_detailed_xlsx(self, request, pk=None):
        batch = self.get_object()
        records = batch.records.select_related("employee", "salary_structure_snapshot").order_by("employee__employee_code")

        columns = [
            ("Employee Name", lambda r: r.employee.full_name),
            ("Email", lambda r: r.employee.email),
            ("Employee Code", lambda r: r.employee.employee_code),
            ("Designation", lambda r: r.employee.position),
            ("CTC", lambda r: r.salary_structure_snapshot.ctc_annual if r.salary_structure_snapshot_id else None),
            ("Comp-Off Opening Balance", lambda r: r.comp_off_opening_balance),
            ("Extra Working Days", lambda r: r.extra_working_days),
            ("Used to Cover LOP", lambda r: r.comp_off_days_used),
            ("Comp-Off Closing Balance", lambda r: r.comp_off_closing_balance),
            ("Paid Leave Days", lambda r: r.paid_leave_days),
            ("LOP Days", lambda r: r.lop_days),
            ("Leave Opening Balance", lambda r: r.leave_opening_balance),
            ("Leave Accrued", lambda r: r.leave_accrued),
            ("Leave Used", lambda r: r.leave_used),
            ("Leave Closing Balance", lambda r: r.leave_closing_balance),
            ("Basic DA", lambda r: r.basic_da),
            ("HRA", lambda r: r.hra),
            ("LTA", lambda r: r.lta),
            ("Special Allowance", lambda r: r.special_allowance),
            ("NPS Allowance Earned", lambda r: r.nps_allowance_earned),
            ("Gross Salary", lambda r: r.gross_salary),
            ("Variable Pay", lambda r: r.variable_pay),
            ("Commission Other", lambda r: r.commission_other),
            ("Arrears", lambda r: r.arrears),
            ("Reimbursements", lambda r: r.reimbursements),
            ("Salary Advance Opening Balance", lambda r: r.salary_advance_opening_balance),
            ("Salary Advance Given", lambda r: r.salary_advance_disbursed),
            ("Salary Advance Balance", lambda r: r.salary_advance_closing_balance),
            ("On Hold Opening Balance", lambda r: r.on_hold_opening_balance),
            ("On Hold Released", lambda r: r.on_hold_released),
            ("Earned Salary", lambda r: r.earned_salary),
            ("EPF", lambda r: r.epf),
            ("VPF", lambda r: r.vpf),
            ("Professional Tax", lambda r: r.professional_tax),
            ("TDS", lambda r: r.tds),
            ("NPS Deduction", lambda r: r.nps_deduction),
            ("Loan Deduction", lambda r: r.loan_deduction),
            ("LWF", lambda r: r.lwf),
            ("Other Deduction", lambda r: r.other_deduction),
            ("Salary Advance Recovered", lambda r: r.salary_advance_recovered),
            ("On Hold Deducted", lambda r: r.on_hold_deducted),
            ("Total Deductions", lambda r: r.total_deductions),
            ("Net Salary", lambda r: r.net_salary),
            ("Status", lambda r: r.get_status_display()),
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "Payroll Batch"

        last_col = len(columns)
        last_col_letter = get_column_letter(last_col)

        # Row 1 — company name
        ws.merge_cells(f"A1:{last_col_letter}1")
        ws["A1"] = batch.client.name
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Row 2 — month/year
        ws.merge_cells(f"A2:{last_col_letter}2")
        ws["A2"] = f"{month_name[batch.month]} {batch.year}"
        ws["A2"].font = Font(bold=True, size=12)
        ws["A2"].alignment = Alignment(horizontal="center")

        # Row 3 — headers (with a merged "Earnings" / "Deductions" group row above them isn't
        # requested; keeping a single header row as specified by the columns list itself)
        header_row = 3
        for col_idx, (label, _) in enumerate(columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=label)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Rows 4+ — data
        for row_idx, record in enumerate(records, start=header_row + 1):
            for col_idx, (_, getter) in enumerate(columns, start=1):
                value = getter(record)
                ws.cell(row=row_idx, column=col_idx, value=float(value) if isinstance(value, Decimal) else value)

        for col_idx in range(1, last_col + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = f"payroll_{batch.client_id}_{batch.year}_{batch.month:02d}_detailed.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=["get"], url_path="records")
    def records(self, request, pk=None):
        batch = self.get_object()
        queryset = batch.records.select_related("employee", "salary_structure_snapshot", "batch", "batch__client")
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = PayslipRecordSerializer(page, many=True, context={"request": request})
            return self.get_paginated_response(serializer.data)
        serializer = PayslipRecordSerializer(queryset, many=True, context={"request": request})
        return Response(serializer.data)

    @action(detail=False, methods=["get"], url_path="stats")
    def stats(self, request):
        """
        Firm-wide payroll snapshot for the PayrollWorkspace overview screen
        (status_overview / pending_review / emails_sent / failed_emails).
        This endpoint didn't exist in the original payroll module — added
        here since the overview page depends on this exact shape.
        """
        batches = PayrollBatch.objects.all()
        status_overview = {
            "uploaded": batches.filter(status=PayrollBatch.STATUS_UPLOADED).count(),
            "reviewed": batches.filter(status=PayrollBatch.STATUS_REVIEWED).count(),
            "in_progress": batches.filter(status=PayrollBatch.STATUS_SENDING).count(),
            "completed": batches.filter(status=PayrollBatch.STATUS_COMPLETED).count(),
            "failed": batches.filter(status=PayrollBatch.STATUS_FAILED).count(),
        }
        pending_review = batches.filter(status=PayrollBatch.STATUS_UPLOADED).count()
        emails_sent = EmailLog.objects.filter(success=True).count()
        failed_emails = EmailLog.objects.filter(success=False).count()

        return Response(
            {
                "status_overview": status_overview,
                "pending_review": pending_review,
                "emails_sent": emails_sent,
                "failed_emails": failed_emails,
            }
        )


class PayslipRecordViewSet(AuditViewMixin, viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsAdminOrManagerRole]
    serializer_class = PayslipRecordSerializer
    http_method_names = ["get", "post", "patch", "head", "options"]

    def get_queryset(self):
        queryset = PayslipRecord.objects.select_related("batch", "batch__client", "employee", "salary_structure_snapshot")
        if batch_id := self.request.query_params.get("batch"):
            queryset = queryset.filter(batch_id=batch_id)
        if status_param := self.request.query_params.get("status"):
            queryset = queryset.filter(status=status_param)
        if employee_id := self.request.query_params.get("employee"):
            queryset = queryset.filter(employee_id=employee_id)
            return queryset.order_by("-batch__year", "-batch__month")
        return queryset.order_by("employee__employee_code")

    def partial_update(self, request, *args, **kwargs):
        record = self.get_object()

        # Once a batch is COMPLETED, its records have already been emailed
        # out as final payslips — editing after the fact would silently
        # desync what the employee received from what CAOAS shows, with no
        # way to re-notify them. Individual EMAIL_SENT records are blocked
        # too, in case a batch is ever partially sent (SENDING with some
        # records already delivered).
        if record.batch.status == PayrollBatch.STATUS_COMPLETED or record.status == PayslipRecord.STATUS_EMAIL_SENT:
            raise PermissionDenied(
                "This payslip has already been emailed and can no longer be edited."
            )

        serializer = self.get_serializer(record, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)

        changed_fields = []
        with transaction.atomic():
            for field_name, new_value in serializer.validated_data.items():
                if field_name not in PAYSLIP_EDITABLE_FIELDS:
                    continue
                old_value = getattr(record, field_name)
                if old_value != new_value:
                    changed_fields.append((field_name, old_value, new_value))

            self.perform_update(serializer)

            # Comp-off/leave ledger AND the proration-driven pay fields
            # (basic_da, hra, epf, professional_tax, earned_salary) are all
            # derived from attendance — if any attendance input is edited
            # post-generation, the stored derived fields are now stale and
            # must be recomputed the same way generation did, via
            # calculate_payslip_fields(), or they silently drift from what
            # the ledger/attendance inputs say they should be (and next
            # month's opening balance would inherit the drift).
            update_fields = ["edit_count", "earned_salary", "total_deductions", "net_salary", "updated_at"]
            changed_field_names = {f for f, _, _ in changed_fields}
            attendance_fields = {"extra_working_days", "lop_days", "actual_working_days", "paid_leave_days"}
            if changed_field_names & attendance_fields:
                # Re-derive is_probation from date_of_joining rather than
                # assuming the record's original probation status still
                # holds (edits can happen well after generation).
                is_probation = is_in_probation(
                    record.employee.hire_date, record.batch.year, record.batch.month
                )
                row = {
                    "actual_working_days": record.actual_working_days,
                    "lop_days": record.lop_days,
                    "extra_working_days": record.extra_working_days,
                    "paid_leave_days": record.paid_leave_days,
                    "lta": record.lta_upload,
                    "special_allowance": record.special_allowance_upload,
                    "nps_allowance_earned": record.nps_allowance_upload,
                    "commission_other": record.commission_other,
                    "arrears": record.arrears,
                }
                computed = calculate_payslip_fields(
                    record.salary_structure_snapshot, row, record.batch.month, record.days_in_month,
                    comp_off_opening_balance=record.comp_off_opening_balance,
                    leave_opening_balance=record.leave_opening_balance,
                    is_probation=is_probation,
                )
                for field_name, value in computed.items():
                    # gross_salary/comp_off_opening_balance/leave_opening_balance
                    # are carried straight through unchanged by design (see
                    # calculate_payslip_fields) — only re-assign the fields
                    # that can actually move from a proration edit.
                    if field_name in ("gross_salary", "comp_off_opening_balance", "leave_opening_balance"):
                        continue
                    setattr(record, field_name, value)
                update_fields += [
                    "basic_da", "hra", "lta", "special_allowance", "nps_allowance_earned",
                    "basic_for_pf", "epf", "vpf", "professional_tax",
                    "comp_off_days_used", "comp_off_closing_balance",
                    "leave_accrued", "leave_used", "leave_closing_balance",
                ]

            if changed_fields:
                PayslipRecordEdit.objects.bulk_create(
                    [
                        PayslipRecordEdit(
                            payslip_record=record,
                            edited_by=request.user,
                            field_name=field_name,
                            old_value=str(old_value),
                            new_value=str(new_value),
                        )
                        for field_name, old_value, new_value in changed_fields
                    ]
                )
                record.edit_count += len(changed_fields)
                record.save(update_fields=update_fields)

        return Response(self.get_serializer(record).data)

    @action(detail=True, methods=["get"], url_path="edit-history")
    def edit_history(self, request, pk=None):
        record = self.get_object()
        serializer = PayslipRecordEditSerializer(record.edits.select_related("edited_by"), many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["post"], url_path="resend")
    def resend(self, request, pk=None):
        record = self.get_object()
        success, message = send_payslip_email(record, request.user)
        return Response({"success": success, "message": message})

    @action(detail=True, methods=["get"], url_path="download-pdf")
    def download_pdf(self, request, pk=None):
        record = self.get_object()
        if not record.pdf_path:
            raise Http404("Payslip PDF has not been generated.")
        pdf_path = Path(record.pdf_path)
        if not pdf_path.exists():
            raise Http404("Payslip PDF file is missing.")
        employee_name_slug = record.employee.full_name.replace(" ", "")
        filename = f"{employee_name_slug}_{month_name[record.batch.month]}{record.batch.year}-payslip.pdf"
        return FileResponse(pdf_path.open("rb"), as_attachment=True, filename=filename, content_type="application/pdf")

    @action(detail=True, methods=["post"], url_path="regenerate-pdf")
    def regenerate_pdf(self, request, pk=None):
        record = self.get_object()
        pdf_path = generate_payslip_pdf(record)
        return Response({"pdf_path": pdf_path, "record": self.get_serializer(record).data})







HEADER_ROW = [
    "Employee Code",
    "Actual Working Days",
    "Extra Working Days",
    "Paid Leave Days",
    "LOP",
    "Leave Travel Allowance",
    "Special Allowance",
    "NPS Allowance",
    "commission/other allowance/Retention Bonus",
    "Arrears",
    "Reimbursements",
    "TDS",
    "VPF Arrears",
    "NPS Deduction - Arrears",
    "Loan Deduction",
    "LWF",
    "Other deduction",
]

SAMPLE_ROW = [
    "EMP001",  # Employee Code
    31,        # Actual Working Days
    0,         # Extra Working Days
    0,         # Paid Leave Days
    0,         # LOP
    0,         # Leave Travel Allowance
    0,         # Special Allowance
    0,         # NPS Allowance
    0,         # commission/other allowance/Retention Bonus
    0,         # Arrears
    0,         # Reimbursements
    0,         # TDS
    0,         # VPF Arrears
    0,         # NPS Deduction - Arrears
    0,         # Loan Deduction
    0,         # LWF
    0,         # Other deduction
]


class PayrollTemplateViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated, IsAdminOrManagerRole]

    @action(detail=False, methods=["get"], url_path="download")
    def download(self, request):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Payroll Upload"

        total_cols = len(HEADER_ROW)
        last_col_letter = get_column_letter(total_cols)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # --- Title rows (1-3), kept purely for readability — the parser
        # doesn't read anything from these. ---
        worksheet.merge_cells(f"A1:{last_col_letter}1")
        worksheet["A1"] = "Client Name"
        worksheet.merge_cells(f"A2:{last_col_letter}2")
        worksheet["A2"] = f"Salary Sheet for the month {timezone.now().strftime('%B-%y')}"
        worksheet.merge_cells(f"A3:{last_col_letter}3")
        worksheet["A3"] = timezone.localtime(timezone.now()).replace(day=1, tzinfo=None)
        worksheet["A3"].number_format = "dd-mmm-yyyy"
        for row in (1, 2, 3):
            worksheet.cell(row=row, column=1).alignment = Alignment(horizontal="center")
            worksheet.cell(row=row, column=1).font = Font(bold=(row != 3))

        # --- Row 4: informational note only — "days in month" is now
        # computed automatically from the batch's month/year, not
        # uploaded, so there's nothing to fill in here. Kept as a row so
        # the header stays on row 5, which is what excel_parser.py reads.
        worksheet.merge_cells(f"A4:{last_col_letter}4")
        worksheet["A4"] = (
            "Days in month, and all salary-structure figures (Basic+DA, HRA, "
            "Gross, EPF, VPF, PT, Earned/Net Salary), are calculated "
            "automatically from each employee's salary structure — do not "
            "add columns for them here."
        )
        worksheet["A4"].font = Font(italic=True, size=9, color="808080")

        # --- Row 5: real header row (this is what excel_parser.py reads) ---
        for col_idx, header in enumerate(HEADER_ROW, start=1):
            cell = worksheet.cell(row=5, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = center

        worksheet.freeze_panes = "A6"

        # --- Row 6: sample data row ---
        for col_idx, value in enumerate(SAMPLE_ROW, start=1):
            worksheet.cell(row=6, column=col_idx, value=value)

        # Column widths
        for col_idx, header in enumerate(HEADER_ROW, start=1):
            width = max(len(word) for word in header.split(" ")) if header else 9
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(len(header) + 2, 9), 30)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="payroll_upload_template.xlsx"'
        return response

    @action(detail=False, methods=["get"], url_path="employee-template")
    def employee_template(self, request):
        """
        Template for the bulk employee-master import.
        Single sheet with: Headers (Row 1) → Sample Employee (Row 2) → Instructions (Rows 4+)
        """
        # Employee master columns, plus one salary-structure input: PF
        # Applicable. Everything else about the structure (Basic/HRA/SA/
        # etc.) is always auto-derived from CTC — see
        # EmployeeSalaryStructure.build_from_ctc in models.py — and only
        # for employees who don't already have a structure on file.
        employee_headers = [
            "Employee Code", "First Name", "Last Name", "Email", "PAN Number",
            "Department", "Position", "Hire Date", "CTC", "PF Applicable",
        ]

        headers = employee_headers

        sample = [
            "EMP001", "Rahul", "Sharma", "rahul.sharma@example.com", "ABCDE1234F",
            "Engineering", "Software Engineer", "2026-01-02", 600000, "yes",
        ]

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Employee Data"
        
        # ── ROW 1: HEADERS ──
        worksheet.append(headers)
        
        # Format headers - bold with blue background
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # ── ROW 2: SAMPLE DATA ──
        worksheet.append(sample)
        for cell in worksheet[2]:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # ── ROW 3: EMPTY (SPACER) ──
        worksheet.append([])
        
        # ── ROW 4 ONWARDS: INSTRUCTIONS ──
        current_row = 4
        
        # Title
        worksheet[f"A{current_row}"] = "INSTRUCTIONS"
        worksheet[f"A{current_row}"].font = Font(bold=True, size=12, color="C00000")
        current_row += 1
        
        # Mandatory Fields Section
        worksheet[f"A{current_row}"] = "MANDATORY FIELDS (Must be filled):"
        worksheet[f"A{current_row}"].font = Font(bold=True, color="C00000")
        current_row += 1
        
        mandatory_fields = [
            ("Employee Code", "Unique identifier for employee (e.g., EMP001, EMP002). Must be unique."),
            ("First Name", "Employee's first name (text, max 100 characters)."),
        ]
        
        for field, description in mandatory_fields:
            worksheet[f"A{current_row}"] = f"  • {field}"
            worksheet[f"A{current_row}"].font = Font(bold=True, size=10)
            worksheet[f"B{current_row}"] = description
            worksheet[f"B{current_row}"].alignment = Alignment(wrap_text=True)
            current_row += 1
        
        # Optional Fields Section
        current_row += 1
        worksheet[f"A{current_row}"] = "OPTIONAL FIELDS (Can be left blank):"
        worksheet[f"A{current_row}"].font = Font(bold=True, color="0070C0")
        current_row += 1
        
        optional_fields = [
            ("Last Name", "Employee's last name (text)."),
            ("Email", "Valid email address (format: name@domain.com)."),
            ("PAN Number", "PAN card number (format: ABCDE1234F - 10 characters)."),
            ("Department", "Department name (text)."),
            ("Position", "Job position/designation (text)."),
            ("Hire Date", "Date of hiring (format: YYYY-MM-DD, e.g., 2024-01-15)."),
            ("CTC", "Cost to Company in annual amount (number, can include decimals). If provided and the employee has no salary structure yet, one is auto-created from this."),
            ("PF Applicable", "Only used when a structure is being auto-created from CTC (see above). Enter 'yes' or 'no', lowercase. Blank = defaults to 'yes'."),
        ]
        
        for field, description in optional_fields:
            worksheet[f"A{current_row}"] = f"  • {field}"
            worksheet[f"A{current_row}"].font = Font(bold=True, size=10)
            worksheet[f"B{current_row}"] = description
            worksheet[f"B{current_row}"].alignment = Alignment(wrap_text=True)
            current_row += 1
        
        # Format Notes
        current_row += 1
        worksheet[f"A{current_row}"] = "FORMAT NOTES:"
        worksheet[f"A{current_row}"].font = Font(bold=True, size=11, color="595959")
        current_row += 1
        
        notes = [
            "Dates must be in YYYY-MM-DD format (e.g., 2024-01-15)",
            "Numbers can be decimals (e.g., 50000.50)",
            "Status values: use lowercase 'active' or 'inactive'",
            "PF Applicable: use lowercase 'yes' or 'no'",
            "Don't include currency symbols (₹, $, etc.)",
            "Existing employees are updated by Employee Code; new codes are created",
            "If CTC is filled and the employee has no salary structure yet, one is auto-created from CTC (Basic/HRA/etc. derived automatically — not entered in this file)",
            "If the employee already has a salary structure, CTC and PF Applicable here only update their Employee record — the structure is not touched",
        ]
        
        for note in notes:
            worksheet[f"A{current_row}"] = f"  • {note}"
            worksheet[f"A{current_row}"].alignment = Alignment(wrap_text=True)
            worksheet[f"A{current_row}"].font = Font(size=10)
            current_row += 1
        
        # Set column widths
        worksheet.column_dimensions["A"].width = 35
        worksheet.column_dimensions["B"].width = 70
        for col_idx in range(3, len(headers) + 1):
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(len(headers[col_idx-1]) + 2, 9), 25)
        
        # Freeze panes at header row
        worksheet.freeze_panes = "A2"
        
        # Format specific columns
        # Hire Date column formatted as a date
        worksheet.cell(row=2, column=employee_headers.index("Hire Date") + 1).number_format = "yyyy-mm-dd"

        # Currency columns
        currency_cols = ["CTC"]
        for col_name in currency_cols:
            if col_name in headers:
                col_idx = headers.index(col_name) + 1
                worksheet.cell(row=2, column=col_idx).number_format = "#,##0.00"

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="employee_master_template.xlsx"'
        return response

















class EmailLogViewSet(viewsets.ReadOnlyModelViewSet):
    # permission_classes = [IsAuthenticated, IsAdminRole]
    permission_classes = [IsAuthenticated, IsAdminOrManagerRole]
    serializer_class = EmailLogSerializer

    def get_queryset(self):
        from .models import EmailLog

        queryset = EmailLog.objects.select_related(
            "payslip_record",
            "payslip_record__employee",
            "payslip_record__batch",
            "payslip_record__batch__client",
            "sent_by",
        )
        if batch_id := self.request.query_params.get("batch"):
            queryset = queryset.filter(payslip_record__batch_id=batch_id)
        if success := self.request.query_params.get("success"):
            if success.lower() in {"true", "1", "sent"}:
                queryset = queryset.filter(success=True)
            elif success.lower() in {"false", "0", "failed"}:
                queryset = queryset.filter(success=False)
        return queryset.order_by("-sent_at")

# ── Ledger adjustments (Comp-Off / Leave) ─────────────────────────────────
# New endpoints, added inside the payroll module only (see note in
# serializers.py above) so BatchReview's ledger-adjustment modal has a
# real API to call. Both are simple create+list; edits/deletes aren't
# exposed since an adjustment is meant to be an immutable ledger entry.

from .serializers import (  # noqa: E402
    CompOffAdjustmentSerializer, LeaveAdjustmentSerializer, OnHoldAdjustmentSerializer,
    SalaryAdvanceAdjustmentSerializer, SalaryAdvanceSerializer,
)


class CompOffAdjustmentViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsAdminOrManagerRole]
    serializer_class = CompOffAdjustmentSerializer
    http_method_names = ["get", "post", "head", "options"]

    def get_queryset(self):
        queryset = CompOffAdjustment.objects.select_related("employee", "created_by").order_by("-created_at")
        if employee_id := self.request.query_params.get("employee"):
            queryset = queryset.filter(employee_id=employee_id)
        return queryset

    def perform_create(self, serializer):
        adjustment = serializer.save(created_by=self.request.user)
        open_record = _current_open_record(adjustment.employee)
        if open_record is not None:
            _apply_comp_off_or_leave_adjustment(open_record, "comp_off", adjustment.amount)
            adjustment.applied_in_record = open_record
            adjustment.save(update_fields=["applied_in_record"])


class LeaveAdjustmentViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsAdminOrManagerRole]
    serializer_class = LeaveAdjustmentSerializer
    http_method_names = ["get", "post", "head", "options"]

    def get_queryset(self):
        queryset = LeaveAdjustment.objects.select_related("employee", "created_by").order_by("-created_at")
        if employee_id := self.request.query_params.get("employee"):
            queryset = queryset.filter(employee_id=employee_id)
        return queryset

    def perform_create(self, serializer):
        adjustment = serializer.save(created_by=self.request.user)
        open_record = _current_open_record(adjustment.employee)
        if open_record is not None:
            _apply_comp_off_or_leave_adjustment(open_record, "leave", adjustment.amount)
            adjustment.applied_in_record = open_record
            adjustment.save(update_fields=["applied_in_record"])


class OnHoldAdjustmentViewSet(viewsets.ModelViewSet):
    """
    Grant ("put on hold", positive amount) / deduct ("release", negative
    amount) manual corrections to an employee's on-hold balance. Same
    pending-until-next-batch pattern as CompOffAdjustment/LeaveAdjustment
    above — see OnHoldAdjustment model docstring.
    """

    permission_classes = [IsAuthenticated, IsAdminOrManagerRole]
    serializer_class = OnHoldAdjustmentSerializer
    http_method_names = ["get", "post", "head", "options"]

    def get_queryset(self):
        queryset = OnHoldAdjustment.objects.select_related("employee", "created_by").order_by("-created_at")
        if employee_id := self.request.query_params.get("employee"):
            queryset = queryset.filter(employee_id=employee_id)
        return queryset

    def perform_create(self, serializer):
        adjustment = serializer.save(created_by=self.request.user)
        open_record = _current_open_record(adjustment.employee)
        if open_record is not None:
            _apply_money_adjustment(open_record, "on_hold", adjustment.amount)
            adjustment.applied_in_record = open_record
            adjustment.save(update_fields=["applied_in_record"])


class SalaryAdvanceAdjustmentViewSet(viewsets.ModelViewSet):
    """
    One-off (non-EMI) manual grant/recovery against an employee's salary
    advance balance — positive amount to give a new advance, negative to
    record a recovery. For an EMI-style plan, use SalaryAdvanceViewSet
    instead, which creates its own disbursement adjustment automatically.
    """

    permission_classes = [IsAuthenticated, IsAdminOrManagerRole]
    serializer_class = SalaryAdvanceAdjustmentSerializer
    http_method_names = ["get", "post", "head", "options"]

    def get_queryset(self):
        queryset = SalaryAdvanceAdjustment.objects.select_related(
            "employee", "created_by", "advance"
        ).order_by("-created_at")
        if employee_id := self.request.query_params.get("employee"):
            queryset = queryset.filter(employee_id=employee_id)
        return queryset

    def perform_create(self, serializer):
        adjustment = serializer.save(created_by=self.request.user)
        open_record = _current_open_record(adjustment.employee)
        if open_record is not None:
            _apply_money_adjustment(open_record, "salary_advance", adjustment.amount)
            adjustment.applied_in_record = open_record
            adjustment.save(update_fields=["applied_in_record"])


class SalaryAdvanceViewSet(viewsets.ModelViewSet):
    """
    EMI-style salary advance plans. Creating one immediately creates its
    lump-sum disbursement SalaryAdvanceAdjustment too (see
    SalaryAdvanceSerializer.create) — the monthly recovery installments
    are generated automatically by PayrollUploadView on each subsequent
    batch, no further manual action needed.
    """

    permission_classes = [IsAuthenticated, IsAdminOrManagerRole]
    serializer_class = SalaryAdvanceSerializer
    http_method_names = ["get", "post", "head", "options"]

    def get_queryset(self):
        queryset = SalaryAdvance.objects.select_related("employee", "created_by").order_by("-created_at")
        if employee_id := self.request.query_params.get("employee"):
            queryset = queryset.filter(employee_id=employee_id)
        return queryset

    def perform_create(self, serializer):
        advance = serializer.save(created_by=self.request.user)
        open_record = _current_open_record(advance.employee)
        if open_record is not None:
            # The plan's lump-sum disbursement (created inside
            # SalaryAdvanceSerializer.create) folds in exactly like a
            # manual one-off grant would — recovery installments still
            # wait for a future batch (_auto_generate_pending_emi_recoveries),
            # only the disbursement itself applies immediately here.
            _apply_money_adjustment(open_record, "salary_advance", advance.disbursement_adjustment.amount)
            advance.disbursement_adjustment.applied_in_record = open_record
            advance.disbursement_adjustment.save(update_fields=["applied_in_record"])


# ── Client profile ─────────────────────────────────────────────────────────

from .serializers import (  # noqa: E402
    ClientSerializer, EmployeeSalaryStructureAutoCalcSerializer, EmployeeSalaryStructureSerializer,
    EmployeeSerializer, EmployeeWithStructureSerializer,
)


class ClientViewSet(viewsets.ModelViewSet):
    """
    One row per payroll-service client of the CA firm. Standard
    list/create/retrieve/update — this replaces the old single-row
    Company model now that payroll is multi-client (every payslip is
    issued on behalf of a Client, not the firm itself).
    """

    permission_classes = [IsAuthenticated, IsAdminOrManagerRole]
    serializer_class = ClientSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    queryset = Client.objects.all().order_by("name")

    @action(detail=True, methods=["get"], url_path="pdf-preview")
    def pdf_preview(self, request, pk=None):
        """
        Renders ONE on-demand preview PDF for this client, using the real
        client (name/logo/address/etc.) with dummy employee/salary data —
        for the "Choose Design" modal, so a design can be judged before
        being saved. Nothing is persisted: the Employee/PayrollBatch/
        PayslipRecord below are built in memory and never saved.
        """
        client = self.get_object()
        try:
            design = int(request.query_params.get("design", client.pdf_design))
        except (TypeError, ValueError):
            raise PermissionDenied("design must be an integer 1-8")
        module = DESIGN_MODULES.get(design, DESIGN_MODULES[1])

        today = date.today()
        dummy_employee = Employee(
            id=0,
            client=client,
            employee_code="EMP001",
            first_name="John",
            last_name="Doe",
            email="john.doe@example.com",
            pan_number="ABCDE1234F",
            department="Operations",
            position="Associate",
            hire_date=date(today.year, 1, 1),
            ctc=Decimal("1200000.00"),
            status=Employee.STATUS_ACTIVE,
        )
        dummy_batch = PayrollBatch(
            id=0,
            client=client,
            month=today.month,
            year=today.year,
            uploaded_by=request.user,
        )
        dummy_record = PayslipRecord(
            id=0,
            batch=dummy_batch,
            employee=dummy_employee,
            days_in_month=30,
            actual_working_days=30,
            paid_leave_days=Decimal("0"),
            lop_days=Decimal("0"),
            basic_da=Decimal("50000.00"),
            hra=Decimal("20000.00"),
            special_allowance=Decimal("15000.00"),
            gross_salary=Decimal("85000.00"),
            earned_salary=Decimal("85000.00"),
            epf=Decimal("1800.00"),
            professional_tax=Decimal("200.00"),
            tds=Decimal("2000.00"),
            total_deductions=Decimal("4000.00"),
            net_salary=Decimal("81000.00"),
        )

        # CTC on the payslip is NOT read from Employee.ctc — every design's
        # _structure() helper looks it up via a real DB query against
        # EmployeeSalaryStructure.objects.filter(employee=record.employee,
        # ...). dummy_employee has no pk, so that query would always return
        # no match and CTC would render as "—". Stub the lookup's queryset
        # chain (.filter().order_by().first()) to hand back this in-memory
        # structure instead, scoped to just this render call.
        dummy_structure = EmployeeSalaryStructure(
            id=0,
            employee=dummy_employee,
            effective_from=date(today.year, 1, 1),
            ctc_annual=dummy_employee.ctc,
            pf_opted=True,
        )

        class _FakeSalaryStructureQuerySet:
            def order_by(self, *args, **kwargs):
                return self

            def first(self):
                return dummy_structure

        pdf_bytes = self._render_preview_pdf(module, dummy_record, _FakeSalaryStructureQuerySet())
        return HttpResponse(
            pdf_bytes,
            content_type="application/pdf",
            headers={"Content-Disposition": "inline; filename=preview.pdf"},
        )

    @staticmethod
    def _render_preview_pdf(module, dummy_record, fake_queryset):
        with patch.object(EmployeeSalaryStructure.objects, "filter", return_value=fake_queryset):
            return module._build_pdf_bytes(dummy_record, encryption=None)


# ── Salary structure ───────────────────────────────────────────────────────

class EmployeeSalaryStructureViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsAdminOrManagerRole]
    serializer_class = EmployeeSalaryStructureSerializer
    http_method_names = ["get", "post", "patch", "delete", "head", "options"]

    def get_queryset(self):
        queryset = EmployeeSalaryStructure.objects.select_related("employee", "created_by")
        if employee_id := self.request.query_params.get("employee"):
            queryset = queryset.filter(employee_id=employee_id)
        return queryset.order_by("-effective_from", "-created_at")

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=["post"], url_path="auto-calc")
    def auto_calc(self, request):
        """
        Given { ctc_annual, pf_opted }, returns the derived structure
        breakdown (Basic/HRA/Special Allowance/etc.) using the firm's
        standard CTC formula, WITHOUT saving anything — the frontend
        shows this as an editable preview before the user confirms.
        """
        body = EmployeeSalaryStructureAutoCalcSerializer(data=request.data)
        body.is_valid(raise_exception=True)
        breakdown = EmployeeSalaryStructure.build_from_ctc(
            body.validated_data["ctc_annual"],
            body.validated_data["pf_opted"],
            body.validated_data["current_lta"],
        )
        return Response({k: str(v) if isinstance(v, Decimal) else v for k, v in breakdown.items()})

    @action(detail=False, methods=["get"], url_path="employees")
    def list_employees(self, request):
        """
        Payroll's own "Employees" list: every payroll.Employee row,
        optionally filtered to one client, alongside their latest salary
        structure, if any. Powers the payroll → Employees screen.
        """
        search = request.query_params.get("search", "").strip()
        qs = Employee.objects.prefetch_related("salary_structures")
        if client_id := request.query_params.get("client"):
            qs = qs.filter(client_id=client_id)
        if search:
            qs = qs.filter(
                models.Q(first_name__icontains=search)
                | models.Q(last_name__icontains=search)
                | models.Q(employee_code__icontains=search)
            )

        rows = []
        for employee in qs.order_by("first_name", "last_name"):
            structure = get_latest_salary_structure(employee)
            rows.append(
                {
                    "id": employee.id,
                    "employee_code": employee.employee_code,
                    "first_name": employee.first_name,
                    "last_name": employee.last_name,
                    "full_name": employee.full_name,
                    "email": employee.email,
                    "pan_number": employee.pan_number,
                    "department": employee.department,
                    "position": employee.position,
                    "hire_date": employee.hire_date,
                    "status": employee.status,
                    "ctc": employee.ctc,
                    "client": employee.client_id,
                    "salary_structure": structure,
                }
            )

        serializer = EmployeeWithStructureSerializer(rows, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"], url_path=r"employees/(?P<employee_id>[^/.]+)")
    def employee_detail(self, request, employee_id=None):
        """Single-employee version of list_employees(), for a detail page."""
        try:
            employee = Employee.objects.get(pk=employee_id)
        except Employee.DoesNotExist:
            raise Http404("Employee not found")

        structure = get_latest_salary_structure(employee)
        row = {
            "id": employee.id,
            "employee_code": employee.employee_code,
            "first_name": employee.first_name,
            "last_name": employee.last_name,
            "full_name": employee.full_name,
            "email": employee.email,
            "pan_number": employee.pan_number,
            "department": employee.department,
            "position": employee.position,
            "hire_date": employee.hire_date,
            "status": employee.status,
            "ctc": employee.ctc,
            "client": employee.client_id,
            "salary_structure": structure,
        }
        return Response(EmployeeWithStructureSerializer(row).data)


# ── Employee master (payroll-local Employee, per client) ──────────────────

class EmployeeViewSet(viewsets.ModelViewSet):
    """
    CRUD for the payroll-local Employee model, plus the bulk Excel import
    action used to populate/update it per client (see
    excel_parser.parse_employee_master_excel). This is deliberately
    separate from the monthly payroll batch upload — this is the
    "employee master" for a client, updated whenever HR-side details
    change; the monthly batch upload only reads against employees already
    on file here.
    """

    permission_classes = [IsAuthenticated, IsAdminOrManagerRole]
    serializer_class = EmployeeSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_queryset(self):
        queryset = Employee.objects.select_related("client").order_by("employee_code")
        if client_id := self.request.query_params.get("client"):
            queryset = queryset.filter(client_id=client_id)
        return queryset

    @action(detail=False, methods=["post"], url_path="import")
    def import_excel(self, request):
        """
        Bulk create/update Employee rows for one client from an uploaded
        Excel file. Body: client_id (or client), file. Rows are matched
        to existing employees by employee_code WITHIN that client —
        create if new, update if the code already exists for this
        client. Bad rows are rejected with a row-level error, never
        silently skipped (same policy as the monthly batch upload).
        """
        from .excel_parser import EmployeeImportError, parse_employee_master_excel

        client_id = request.data.get("client_id") or request.data.get("client")
        file = request.FILES.get("file")
        if not client_id or not file:
            return Response({"detail": "client_id and file are required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            client = Client.objects.get(pk=client_id)
        except (Client.DoesNotExist, ValueError, TypeError):
            return Response({"detail": "client_id does not match a known client."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            created_count, updated_count, warnings = parse_employee_master_excel(
                file, client, created_by=request.user
            )
        except EmployeeImportError as exc:
            return Response(
                {
                    "detail": f"{len(exc.errors)} rows have errors",
                    "errors": exc.errors,
                    "total_errors": len(exc.errors),
                },
                status=status.HTTP_422_UNPROCESSABLE_ENTITY,
            )

        return Response(
            {"created": created_count, "updated": updated_count, "warnings": warnings},
            status=status.HTTP_200_OK,
        )